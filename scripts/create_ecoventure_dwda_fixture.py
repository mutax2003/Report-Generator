#!/usr/bin/env python3
"""Create minimal Ecoventure workbook fixture for CI tests."""

from __future__ import annotations

import sys
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parents[1]
OUT = ROOT / "samples" / "ecoventure_dwda" / "minimal_calc_workbook.xlsx"


def main() -> int:
    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    p1 = wb.create_sheet("Phase 1 Data")
    p1["B4"] = "Authorization Holder\n (Client Name)"
    p1["T4"] = "Well name or UWI"
    p1["U4"] = "Spud Date"
    p1["V4"] = "Final Drill Date"
    p1["W4"] = "Well Depth (m)"
    p1["AA4"] = "Volume (m3)"
    p1["AB4"] = "Disposal Method"
    p1["AC4"] = "Sump Type"
    p1["AE4"] = "Disposal Location(s)"
    p1["AG4"] = "Drilling Waste Disposal -  [Select all that apply]"
    p1["AI4"] = "Are there any other drilling waste comments you wish to add?"
    p1["B5"] = "Fixture Energy Ltd."
    p1["T5"] = "FIXTURE 1-1-1-1W1M"
    p1["U5"] = "2020-01-01"
    p1["V5"] = "2020-01-05"
    p1["W5"] = 500
    p1["AA5"] = 12
    p1["AB5"] = "LWD"
    p1["AC5"] = "on-lease sump"
    p1["AE5"] = "well centre"
    p1["AG5"] = "Option 1"
    p1["AI5"] = "Fixture disposal summary"

    metal = wb.create_sheet("Metal Calcs (Options 1 &2)")
    metal["B20"] = 40
    metal["E20"] = 500
    metal["G20"] = 3
    metal["I20"] = 40 / (500 * 3)

    salt = wb.create_sheet("Salt Calculations (Option 2)")
    salt["D20"] = 10
    salt["F20"] = 1
    salt["H20"] = 10
    salt["H48"] = 10
    salt["H49"] = 100
    salt["H50"] = 0.1

    dst = wb.create_sheet("Drill Stem Test Returns")
    vol = (76 / 2000) ** 2 * 3.14 * 10
    sacks = vol * 0.28 / 5
    dst["B22"] = 76
    dst["C22"] = 10
    dst["D22"] = vol
    dst["F22"] = 0.28
    dst["H22"] = 5
    dst["J22"] = sacks
    dst["J28"] = sacks

    wb.save(OUT)
    print(f"Wrote {OUT}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
