"""
Ingest Ecoventure Phase I + DWDA calculation workbook (.xlsx saved from xltm).
"""

from __future__ import annotations

import io
import json
import re
from functools import lru_cache
from pathlib import Path
from typing import Any

import openpyxl
import pandas as pd

from compliance_helpers import norm_key

_CONTRACT_PATH = (
    Path(__file__).resolve().parent / "schemas" / "ecoventure_dwda_cell_contract.json"
)
ECOVENTURE_DWDA_TEMPLATE_DIR = (
    Path(__file__).resolve().parent / "templates" / "ecoventure_dwda"
)
ECOVENTURE_FOLDER_FILENAME = "ecoventure_workbook.xlsx"
ENGINE_SUPPORTED_CONTRACT_VERSION = "1.0.0"
DWDA_CALC_SHEET_SKIP_KEYS = frozenset(
    {"calc_type", "notes", "input_json", "result_value", "objective", "pass"}
)
_QP_TEMPLATE_MAP = {
    "2025 Phase 1 ESA - Excel Sheets_TEMPLATE (1).xltm": "ecoventure_phase1_workbook.xltm",
    "25XXXX_DWDA Compliance Option Form (1).dotm": "dwda_compliance_option_form.dotm",
    "25XXXXR Phase 1 ESA Letter_Template (1).dotm": "phase1_esa_letter_template.dotm",
}


def _norm_header(text: str) -> str:
    return re.sub(r"\s+", " ", str(text or "").strip().lower())


@lru_cache(maxsize=1)
def load_cell_contract() -> dict[str, Any]:
    return json.loads(_CONTRACT_PATH.read_text(encoding="utf-8"))


@lru_cache(maxsize=1)
def _signature_sheets() -> frozenset[str]:
    return frozenset(load_cell_contract().get("workbook_signature_sheets", []))


@lru_cache(maxsize=1)
def _contract_metadata() -> dict[str, str]:
    contract = load_cell_contract()
    return {
        "contract_version": str(contract.get("contract_version", "") or ""),
        "workbook_template_id": str(contract.get("workbook_template_id", "") or ""),
    }


def get_cell_contract_version() -> str:
    return _contract_metadata()["contract_version"]


def get_workbook_template_id() -> str:
    return _contract_metadata()["workbook_template_id"]


def cell_contract_provenance() -> dict[str, str]:
    return dict(_contract_metadata())


def contract_ingest_warnings() -> list[str]:
    ver = get_cell_contract_version()
    if ver and ver != ENGINE_SUPPORTED_CONTRACT_VERSION:
        return [
            f"Cell contract version {ver} may not match engine support "
            f"({ENGINE_SUPPORTED_CONTRACT_VERSION})."
        ]
    return []


def maybe_merge_ecoventure_from_folder(
    excel_bytes: bytes,
    folder_root: Path,
) -> tuple[bytes, list[str]]:
    """Merge optional ecoventure_workbook.xlsx from a project folder root."""
    eco_path = folder_root / ECOVENTURE_FOLDER_FILENAME
    if not eco_path.is_file():
        return excel_bytes, []
    warnings = contract_ingest_warnings()
    try:
        return merge_into_engine_excel(excel_bytes, eco_path.read_bytes()), warnings
    except ValueError as e:
        return excel_bytes, [*warnings, f"Ecoventure workbook not merged: {e}"]


def is_ecoventure_workbook(
    path_or_bytes: str | Path | bytes,
    *,
    sheetnames: frozenset[str] | set[str] | None = None,
) -> bool:
    """True when workbook contains signature Ecoventure sheets."""
    required = _signature_sheets()
    if sheetnames is not None:
        return required.issubset(sheetnames)
    try:
        if isinstance(path_or_bytes, (str, Path)):
            wb = openpyxl.load_workbook(path_or_bytes, read_only=True, data_only=True)
        else:
            wb = openpyxl.load_workbook(
                io.BytesIO(path_or_bytes), read_only=True, data_only=True
            )
        ok = required.issubset(wb.sheetnames)
        wb.close()
        return ok
    except Exception:
        return False


@lru_cache(maxsize=1)
def _phase1_data_mapping() -> dict[str, Any]:
    return load_cell_contract().get("phase1_data_mapping", {})


@lru_cache(maxsize=1)
def _calculation_output_specs() -> tuple[tuple[str, dict[str, Any]], ...]:
    return tuple(load_cell_contract().get("calculation_outputs", {}).items())


@lru_cache(maxsize=1)
def _phase1_row_layout() -> tuple[int, int]:
    pmap = _phase1_data_mapping()
    return int(pmap.get("header_row", 4)), int(pmap.get("data_row", 5))


def extract_dwda_calc_outputs(wb: Any) -> dict[str, Any]:
    """Read calculation output cells per cell contract."""
    outputs: dict[str, Any] = {}
    sheetnames = set(wb.sheetnames)
    worksheets: dict[str, Any] = {}
    for key, spec in _calculation_output_specs():
        if "constant" in spec:
            outputs[key] = spec["constant"]
            continue
        sheet = spec.get("sheet")
        cell = spec.get("cell")
        if not sheet or not cell or sheet not in sheetnames:
            continue
        ws = worksheets.get(sheet)
        if ws is None:
            ws = wb[sheet]
            worksheets[sheet] = ws
        val = ws[cell].value
        if val is not None and str(val).strip():
            outputs[key] = val
    return outputs


def _header_map(ws: Any, header_row: int) -> dict[str, int]:
    mapping: dict[str, int] = {}
    for col in range(1, ws.max_column + 1):
        val = ws.cell(row=header_row, column=col).value
        if val:
            mapping[_norm_header(str(val))] = col
    return mapping


@lru_cache(maxsize=1)
def _phase1_column_specs() -> tuple[tuple[str, str], ...]:
    return tuple(
        (field, _norm_header(header))
        for field, header in (_phase1_data_mapping().get("columns") or {}).items()
    )


@lru_cache(maxsize=1)
def _phase1_waste_column_specs() -> tuple[tuple[str, str], ...]:
    return tuple(
        (field, _norm_header(header))
        for field, header in (
            _phase1_data_mapping().get("drilling_waste_columns") or {}
        ).items()
    )


def _mapped_row_values(
    ws: Any,
    headers: dict[str, int],
    data_row: int,
    field_specs: tuple[tuple[str, str], ...],
) -> dict[str, Any]:
    out: dict[str, Any] = {}
    for field, norm_header in field_specs:
        col = headers.get(norm_header)
        if not col:
            continue
        val = ws.cell(row=data_row, column=col).value
        if val is not None and str(val).strip():
            out[field] = str(val).strip()
    return out


def _extract_phase1_data(wb: Any) -> tuple[dict[str, Any], list[dict[str, Any]]]:
    """Map Phase 1 Data row to ProjectData fields and optional DrillingWaste row."""
    header_row, data_row = _phase1_row_layout()
    if "Phase 1 Data" not in wb.sheetnames:
        return {}, []
    ws = wb["Phase 1 Data"]
    headers = _header_map(ws, header_row)
    project = _mapped_row_values(ws, headers, data_row, _phase1_column_specs())
    if project.get("well_name") and not project.get("uwi"):
        project["uwi"] = project["well_name"]

    waste_row = _mapped_row_values(
        ws, headers, data_row, _phase1_waste_column_specs()
    )
    waste_rows: list[dict[str, Any]] = []
    if waste_row and (
        waste_row.get("volume_m3")
        or waste_row.get("disposal_method")
        or waste_row.get("location")
    ):
        waste_row.setdefault("disposal_type", "on-lease")
        waste_rows.append(waste_row)
    return project, waste_rows


def extract_phase1_row(wb: Any) -> dict[str, Any]:
    """Map Ecoventure Phase 1 Data row to ProjectData snake_case keys."""
    project, _ = _extract_phase1_data(wb)
    return project


def extract_drilling_waste_rows(wb: Any) -> list[dict[str, Any]]:
    """Build DrillingWaste row(s) from Phase 1 Data disposal columns."""
    _, waste_rows = _extract_phase1_data(wb)
    return waste_rows


def extract_ecoventure_workbook(workbook_bytes: bytes) -> dict[str, Any]:
    """Full extract: ProjectData fields, DrillingWaste, calc outputs."""
    wb = openpyxl.load_workbook(io.BytesIO(workbook_bytes), data_only=True)
    try:
        names = frozenset(wb.sheetnames)
        if not _signature_sheets().issubset(names):
            raise ValueError("Not an Ecoventure Phase I workbook (missing signature sheets)")
        project, waste = _extract_phase1_data(wb)
        prov = cell_contract_provenance()
        return {
            "project_data": project,
            "drilling_waste": waste,
            "calc_outputs": extract_dwda_calc_outputs(wb),
            "contract_provenance": prov,
            "ingest_warnings": contract_ingest_warnings(),
        }
    finally:
        wb.close()


def flat_calc_row_from_sheet_record(row: dict[str, Any]) -> dict[str, Any]:
    """Normalize DwdaCalculations sheet row to ingest keys."""
    out: dict[str, Any] = {}
    for k, v in row.items():
        nk = norm_key(str(k))
        if nk in DWDA_CALC_SHEET_SKIP_KEYS or v is None or (
            isinstance(v, float) and pd.isna(v)
        ):
            continue
        if str(v).strip():
            out[nk] = v
    return out


def _dwda_calculations_records(calc_outputs: dict[str, Any]) -> list[dict[str, str]]:
    if not calc_outputs:
        return []
    row: dict[str, str] = {
        "calc_type": "ecoventure_ingest",
        "notes": "From Ecoventure workbook",
    }
    for k, v in calc_outputs.items():
        row[k] = str(v) if v is not None else ""
    return [row]


def _merge_project_data(
    project_df: pd.DataFrame,
    merge_data: dict[str, Any],
    project_row_index: int,
) -> pd.DataFrame:
    if not merge_data:
        return project_df
    idx = min(project_row_index, len(project_df) - 1) if len(project_df) else 0
    if project_df.empty:
        return pd.DataFrame({k: str(v) for k, v in merge_data.items()}, index=[0])
    for col, val in merge_data.items():
        if col not in project_df.columns:
            project_df[col] = None
        project_df[col] = project_df[col].astype(object)
        project_df.at[idx, col] = str(val)
    return project_df


def _append_drilling_waste(
    sheets: dict[str, pd.DataFrame],
    waste_rows: list[dict[str, Any]],
) -> None:
    if not waste_rows:
        return
    dw_df = pd.DataFrame(waste_rows)
    existing = sheets.get("DrillingWaste")
    if existing is not None and not existing.empty:
        for col in dw_df.columns:
            if col not in existing.columns:
                existing[col] = ""
        for col in existing.columns:
            if col not in dw_df.columns:
                dw_df[col] = ""
        sheets["DrillingWaste"] = pd.concat([existing, dw_df], ignore_index=True)
    else:
        sheets["DrillingWaste"] = dw_df


def _excel_bytes_from_sheets(sheets: dict[str, pd.DataFrame]) -> bytes:
    bio_out = io.BytesIO()
    with pd.ExcelWriter(bio_out, engine="openpyxl") as writer:
        for name, df in sheets.items():
            df.to_excel(writer, sheet_name=name, index=False)
    return bio_out.getvalue()


def merge_into_engine_excel(
    existing_bytes: bytes,
    ecoventure_bytes: bytes,
    *,
    project_row_index: int = 0,
) -> bytes:
    """
    Merge Ecoventure workbook data into standard engine Excel bytes.
    Overwrites ProjectData row, appends/replaces DrillingWaste, adds DwdaCalculations.
    """
    extracted = extract_ecoventure_workbook(ecoventure_bytes)
    sheets: dict[str, pd.DataFrame] = pd.read_excel(
        io.BytesIO(existing_bytes), sheet_name=None, engine="openpyxl"
    )

    pd_name = "ProjectData"
    if pd_name not in sheets:
        raise ValueError(f"Target Excel missing sheet '{pd_name}'")

    sheets[pd_name] = _merge_project_data(
        sheets[pd_name],
        extracted.get("project_data") or {},
        project_row_index,
    )

    waste_rows = extracted.get("drilling_waste") or []
    _append_drilling_waste(sheets, waste_rows)

    calc_records = _dwda_calculations_records(extracted.get("calc_outputs") or {})
    if calc_records:
        sheets["DwdaCalculations"] = pd.DataFrame(calc_records)

    return _excel_bytes_from_sheets(sheets)


@lru_cache(maxsize=1)
def list_qp_template_files() -> tuple[tuple[str, Path], ...]:
    """Return (zip_name, path) for QP templates in deliverable zip."""
    if not ECOVENTURE_DWDA_TEMPLATE_DIR.is_dir():
        return ()
    out: list[tuple[str, Path]] = []
    for src_name, zip_name in _QP_TEMPLATE_MAP.items():
        path = ECOVENTURE_DWDA_TEMPLATE_DIR / src_name
        if path.is_file():
            out.append((zip_name, path))
    return tuple(out)


@lru_cache(maxsize=8)
def read_qp_template_bytes(path: str) -> bytes:
    """Cached read for Streamlit download buttons."""
    return Path(path).read_bytes()
