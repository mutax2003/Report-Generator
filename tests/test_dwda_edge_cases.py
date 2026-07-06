"""DWDA / Ecoventure integration — edge cases and boundary conditions."""

from __future__ import annotations

import io
import unittest
import zipfile
from pathlib import Path

import openpyxl
import pandas as pd

from compliance_helpers import parse_float
from dwda_calculations import (
    DwdaCalcResult,
    apply_dwda_calc_to_context,
    calc_dst_resistivity_sacks,
    calc_metal_sacks_per_metre,
    calc_salt_max_sacks_per_m3,
    calc_salt_sacks_per_m3,
    evaluate_dwda_calculations,
)
from dwda_compliance import (
    LWD_CUTTINGS_THRESHOLD_M3,
    derive_cuttings_volume_on_lease_m3,
    determine_checklist_scope,
    enrich_dwda_context,
    evaluate_dwda_compliance,
    normalize_compliance_option,
)
from ecoventure_workbook import (
    DWDA_CALC_SHEET_SKIP_KEYS,
    extract_ecoventure_workbook,
    flat_calc_row_from_sheet_record,
    is_ecoventure_workbook,
    merge_into_engine_excel,
)
from engine import ReportEngine, _merge_dwda_calc_sheet

ROOT = Path(__file__).resolve().parents[1]
BASE = ROOT / "samples" / "phase1_alberta_data.xlsx"
TEMPLATE = ROOT / "samples" / "phase1_alberta_template.docx"
FIXTURE = ROOT / "samples" / "ecoventure_dwda" / "minimal_calc_workbook.xlsx"

_SIGNATURE = (
    "Phase 1 Data",
    "Metal Calcs (Options 1 &2)",
    "Salt Calculations (Option 2)",
    "Drill Stem Test Returns",
)


def _build_ecoventure_xlsx(
    *,
    include_phase1: bool = True,
    metal_i20: float | None = 0.027,
    salt_h50: float | None = 0.1,
    dst_j28: float | None = 1.5,
    client: str = "Edge Case Client",
) -> bytes:
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    if include_phase1:
        p1 = wb.create_sheet("Phase 1 Data")
        p1["B4"] = "Authorization Holder\n (Client Name)"
        p1["W4"] = "Well Depth (m)"
        p1["AA4"] = "Volume (m3)"
        p1["AB4"] = "Disposal Method"
        p1["AE4"] = "Disposal Location(s)"
        p1["B5"] = client
        p1["W5"] = 400
        p1["AA5"] = 5
        p1["AB5"] = "LWD"
        p1["AE5"] = "well centre"
    else:
        wb.create_sheet("Phase 1 Data")
    metal = wb.create_sheet("Metal Calcs (Options 1 &2)")
    if metal_i20 is not None:
        metal["I20"] = metal_i20
    salt = wb.create_sheet("Salt Calculations (Option 2)")
    if salt_h50 is not None:
        salt["H48"] = 8
        salt["H49"] = 80
        salt["H50"] = salt_h50
    dst = wb.create_sheet("Drill Stem Test Returns")
    if dst_j28 is not None:
        dst["J28"] = dst_j28
    bio = io.BytesIO()
    wb.save(bio)
    return bio.getvalue()


class ParseFloatEdgeTests(unittest.TestCase):
    def test_none_and_blank(self) -> None:
        self.assertIsNone(parse_float(None))
        self.assertIsNone(parse_float(""))
        self.assertIsNone(parse_float("   "))

    def test_excel_errors(self) -> None:
        self.assertIsNone(parse_float("#DIV/0!"))
        self.assertIsNone(parse_float("#VALUE!"))
        self.assertIsNone(parse_float("nan"))
        self.assertIsNone(parse_float("NaN"))

    def test_commas_and_invalid(self) -> None:
        self.assertEqual(parse_float("1,234.5"), 1234.5)
        self.assertIsNone(parse_float("not-a-number"))
        self.assertEqual(parse_float(42), 42.0)


class CalcBoundaryTests(unittest.TestCase):
    def test_metal_invalid_denominators(self) -> None:
        self.assertIsNone(calc_metal_sacks_per_metre(10, 0, 3))
        self.assertIsNone(calc_metal_sacks_per_metre(10, 100, 0))

    def test_metal_exactly_at_objective_passes(self) -> None:
        # 0.22 = 330 / (500 * 3)
        spm = calc_metal_sacks_per_metre(330, 500, 3)
        assert spm is not None
        self.assertAlmostEqual(spm, 0.22)
        result = evaluate_dwda_calculations(
            {"_ecoventure_ingested": {"metal_sacks_per_metre": spm}}
        )
        self.assertTrue(result.metal_pass)

    def test_metal_just_over_objective_fails(self) -> None:
        result = evaluate_dwda_calculations(
            {"_ecoventure_ingested": {"metal_sacks_per_metre": 0.2201}}
        )
        self.assertFalse(result.metal_pass)
        self.assertTrue(result.phase2_required)

    def test_salt_zero_volume(self) -> None:
        self.assertIsNone(calc_salt_sacks_per_m3(10, 0))

    def test_salt_max_zero_depth(self) -> None:
        self.assertIsNone(calc_salt_max_sacks_per_m3(0))

    def test_salt_at_max_boundary_passes(self) -> None:
        depth = 500.0
        max_spm = calc_salt_max_sacks_per_m3(depth)
        assert max_spm is not None
        result = evaluate_dwda_calculations(
            {
                "well_depth_m": depth,
                "_ecoventure_ingested": {
                    "salt_sacks_per_m3": max_spm,
                    "salt_naoh_equiv_total": max_spm * 100,
                    "salt_waste_volume_m3": 100,
                },
            }
        )
        self.assertTrue(result.salt_pass)

    def test_salt_over_max_fails(self) -> None:
        result = evaluate_dwda_calculations(
            {
                "well_depth_m": 100,
                "_ecoventure_ingested": {
                    "salt_sacks_per_m3": 5.0,
                    "salt_naoh_equiv_total": 500,
                    "salt_waste_volume_m3": 100,
                },
            }
        )
        self.assertFalse(result.salt_pass)
        self.assertTrue(any("Salt" in r for r in result.phase2_reasons))

    def test_dst_resistivity_zero_ohms(self) -> None:
        self.assertIsNone(calc_dst_resistivity_sacks(1.0, 0.28, 0))

    def test_empty_context_summary(self) -> None:
        result = evaluate_dwda_calculations({})
        self.assertEqual(result.summary, "No DWDA calculation inputs present.")
        self.assertFalse(result.phase2_required)
        self.assertEqual(result.to_context_dict()["dwda_metal_pass"], "")

    def test_cross_check_warns_on_mismatch(self) -> None:
        ctx = {
            "well_depth_m": 500,
            "_ecoventure_ingested": {
                "metal_barite_sacks": 400,
                "metal_well_depth_m": 500,
                "metal_mix_ratio": 3,
                "metal_sacks_per_metre": 0.01,
            },
        }
        result = evaluate_dwda_calculations(ctx)
        self.assertTrue(result.cross_check_warnings)
        self.assertTrue(
            any("Metal sacks/m" in w for w in result.cross_check_warnings)
        )

    def test_cross_check_silent_when_aligned(self) -> None:
        spm = calc_metal_sacks_per_metre(40, 500, 3)
        ctx = {
            "well_depth_m": 500,
            "_ecoventure_ingested": {
                "metal_barite_sacks": 40,
                "metal_well_depth_m": 500,
                "metal_mix_ratio": 3,
                "metal_sacks_per_metre": spm,
            },
        }
        result = evaluate_dwda_calculations(ctx)
        self.assertFalse(result.cross_check_warnings)

    def test_dst_rows_from_context(self) -> None:
        ctx = {
            "dst_returns": [
                {
                    "pipe_id_mm": 76,
                    "return_length_m": 10,
                    "resistivity_ohms": 5,
                    "resistivity_factor": 0.28,
                }
            ],
        }
        result = evaluate_dwda_calculations(ctx)
        self.assertIsNotNone(result.dst_resistivity_sacks_total)
        self.assertTrue(result.dst_pass)

    def test_option_2_without_dst_triggers_reason(self) -> None:
        result = evaluate_dwda_calculations(
            {"aer_waste_compliance_option": "Option 2"},
            compliance_option="option_2",
        )
        self.assertTrue(
            any("DST" in r for r in result.phase2_reasons)
        )

    def test_option_1_without_dst_no_dst_reason(self) -> None:
        result = evaluate_dwda_calculations(
            {"aer_waste_compliance_option": "Option 1"},
            compliance_option="option_1",
        )
        self.assertFalse(any("DST" in r for r in result.phase2_reasons))

    def test_dwda_calculations_loop_ingest(self) -> None:
        ctx = {
            "dwda_calculations": [
                {
                    "calc_type": "metal_sacks_per_metre",
                    "result_value": "0.05",
                }
            ],
        }
        result = evaluate_dwda_calculations(ctx)
        self.assertAlmostEqual(result.metal_sacks_per_metre or 0, 0.05)

    def test_apply_dwda_calc_preserves_context_keys(self) -> None:
        ctx = {"well_name": "Test", "client_name": "Co"}
        out = apply_dwda_calc_to_context(ctx)
        self.assertEqual(out["well_name"], "Test")
        self.assertIn("_dwda_calc_result", out)


class ComplianceScopeEdgeTests(unittest.TestCase):
    def test_normalize_variants(self) -> None:
        self.assertEqual(normalize_compliance_option(""), "")
        self.assertEqual(normalize_compliance_option(None), "")
        self.assertEqual(normalize_compliance_option("Approved Facility"), "approved_facility")
        self.assertEqual(normalize_compliance_option("Option 2"), "option_2")
        self.assertEqual(normalize_compliance_option("option_3"), "option_3")
        self.assertEqual(normalize_compliance_option("No on-site waste"), "no_on_site_waste")
        self.assertEqual(normalize_compliance_option("custom path"), "custom_path")

    def test_cuttings_exactly_50_not_full(self) -> None:
        ctx = {
            "aer_waste_compliance_option": "Option 1",
            "cuttings_volume_on_lease_m3": str(LWD_CUTTINGS_THRESHOLD_M3),
            "drilling_waste": [
                {
                    "disposal_method": "remote haul",
                    "disposal_type": "off-lease",
                    "location": "remote",
                }
            ],
        }
        _, scope, cuttings = determine_checklist_scope(ctx)
        self.assertEqual(cuttings, 50.0)
        self.assertEqual(scope, "option_1_minimal")

    def test_cuttings_just_over_50_full(self) -> None:
        ctx = {
            "aer_waste_compliance_option": "Option 1",
            "cuttings_volume_on_lease_m3": "50.1",
            "drilling_waste": [],
        }
        _, scope, _ = determine_checklist_scope(ctx)
        self.assertEqual(scope, "option_1_full")

    def test_no_on_site_waste_scope_none(self) -> None:
        ctx = {
            "no_drilling_waste_on_site": "Yes",
            "drilling_waste": [],
        }
        option, scope, cuttings = determine_checklist_scope(ctx)
        self.assertEqual(scope, "none")
        self.assertIsNone(cuttings)

    def test_option_2_scope_fixed(self) -> None:
        _, scope, _ = determine_checklist_scope(
            {"aer_waste_compliance_option": "Option 2", "cuttings_volume_on_lease_m3": "5"}
        )
        self.assertEqual(scope, "option_2")

    def test_derive_cuttings_no_on_lease_volumes(self) -> None:
        rows = [{"volume_m3": "20", "disposal_type": "off-lease", "location": "remote"}]
        self.assertIsNone(derive_cuttings_volume_on_lease_m3(rows))

    def test_derive_cuttings_ignores_non_dict(self) -> None:
        rows = ["bad", {"volume_m3": "10", "disposal_type": "on-lease", "location": "x"}]
        self.assertEqual(derive_cuttings_volume_on_lease_m3(rows), 10.0)

    def test_evaluate_returns_none_for_non_phase1_profile(self) -> None:
        self.assertIsNone(
            evaluate_dwda_compliance({}, {}, report_type="phase2_esa")
        )

    def test_enrich_skips_non_dwda_profile(self) -> None:
        ctx = {"_report_type": "phase2_esa", "well_name": "X"}
        out = enrich_dwda_context(ctx, {})
        self.assertNotIn("dwda_compliance_summary", out)

    def test_enrich_calc_failure_sets_phase2(self) -> None:
        ctx = {
            "_report_type": "phase1_alberta",
            "aer_waste_compliance_option": "Option 1",
            "drilling_waste_summary": "S",
            "directive_050_notification_ref": "N-1",
            "drilling_waste": [{"disposal_method": "LWD", "location": "on lease", "gps_coordinates": "x"}],
            "_ecoventure_ingested": {"metal_sacks_per_metre": 1.0},
        }
        out = enrich_dwda_context(ctx, {})
        self.assertEqual(out.get("dwda_metal_pass"), "No")
        self.assertEqual(out.get("dwda_calc_phase2_required"), "Yes")


class EcoventureIngestEdgeTests(unittest.TestCase):
    @classmethod
    def setUpClass(cls) -> None:
        if not FIXTURE.is_file():
            import subprocess
            import sys

            subprocess.check_call(
                [sys.executable, str(ROOT / "scripts" / "create_ecoventure_dwda_fixture.py")]
            )

    def test_garbage_bytes_not_ecoventure(self) -> None:
        self.assertFalse(is_ecoventure_workbook(b"not an xlsx"))

    def test_partial_sheets_via_sheetnames(self) -> None:
        self.assertFalse(
            is_ecoventure_workbook(b"x", sheetnames=frozenset(["Phase 1 Data"]))
        )
        self.assertTrue(
            is_ecoventure_workbook(b"x", sheetnames=frozenset(_SIGNATURE))
        )

    def test_extract_rejects_non_ecoventure(self) -> None:
        bio = io.BytesIO()
        wb = openpyxl.Workbook()
        wb.save(bio)
        with self.assertRaises(ValueError):
            extract_ecoventure_workbook(bio.getvalue())

    def test_empty_phase1_row(self) -> None:
        data = extract_ecoventure_workbook(_build_ecoventure_xlsx(include_phase1=False))
        self.assertEqual(data["project_data"], {})
        self.assertEqual(data["drilling_waste"], [])

    def test_merge_missing_projectdata_raises(self) -> None:
        bio = io.BytesIO()
        with pd.ExcelWriter(bio, engine="openpyxl") as w:
            pd.DataFrame({"x": [1]}).to_excel(w, sheet_name="Other", index=False)
        with self.assertRaises(ValueError):
            merge_into_engine_excel(bio.getvalue(), _build_ecoventure_xlsx())

    def test_flat_calc_row_skips_metadata(self) -> None:
        row = {
            "calc_type": "ecoventure_ingest",
            "notes": "n",
            "metal_sacks_per_metre": 0.1,
            "result_value": 99,
            "pass": "Yes",
        }
        flat = flat_calc_row_from_sheet_record(row)
        self.assertEqual(flat.get("metal_sacks_per_metre"), 0.1)
        for k in DWDA_CALC_SHEET_SKIP_KEYS:
            self.assertNotIn(k, flat)

    def test_extract_includes_metal_objective_constant(self) -> None:
        data = extract_ecoventure_workbook(_build_ecoventure_xlsx())
        self.assertEqual(data["calc_outputs"].get("metal_objective"), 0.22)

    def test_merge_dwda_calc_sheet_engine_helper(self) -> None:
        ctx: dict = {
            "dwda_calc_sheet": [
                {
                    "calc_type": "ecoventure_ingest",
                    "metal_sacks_per_metre": 0.15,
                    "notes": "x",
                }
            ],
            "well_name": "Keep",
        }
        _merge_dwda_calc_sheet(ctx)
        self.assertNotIn("dwda_calc_sheet", ctx)
        self.assertEqual(ctx["metal_sacks_per_metre"], 0.15)
        self.assertIn("_ecoventure_ingested", ctx)
        self.assertEqual(ctx["well_name"], "Keep")

    def test_merge_dwda_calc_sheet_empty_noop(self) -> None:
        ctx: dict = {"dwda_calc_sheet": []}
        _merge_dwda_calc_sheet(ctx)
        self.assertNotIn("_ecoventure_ingested", ctx)


class DwdaRenderEdgeTests(unittest.TestCase):
    @classmethod
    def setUpClass(cls) -> None:
        if not BASE.is_file() or not TEMPLATE.is_file():
            raise unittest.SkipTest("samples missing")
        if not FIXTURE.is_file():
            import subprocess
            import sys

            subprocess.check_call(
                [sys.executable, str(ROOT / "scripts" / "create_ecoventure_dwda_fixture.py")]
            )

    def test_render_with_merged_ecoventure_excel(self) -> None:
        merged = merge_into_engine_excel(BASE.read_bytes(), FIXTURE.read_bytes())
        engine = ReportEngine(merged, TEMPLATE.read_bytes())
        _, _, ctx, _ = engine.render(
            meta={"prepared_by": "Edge", "date_of_issue": "2026-06-10", "report_phase": "Phase 1"},
            excel_filename="merged.xlsx",
            template_filename=TEMPLATE.name,
        )
        self.assertIn("dwda_calc_summary", ctx)
        self.assertIn("dwda_calculations", ctx)

    def test_deliverable_zip_has_qp_templates(self) -> None:
        from deliverable_pack import AppendixFile, build_deliverable_zip_bytes

        merged = merge_into_engine_excel(BASE.read_bytes(), FIXTURE.read_bytes())
        engine = ReportEngine(merged, TEMPLATE.read_bytes())
        docx, _, ctx, record = engine.render(
            meta={"prepared_by": "Edge", "date_of_issue": "2026-06-10", "report_phase": "Phase 1"},
            appendix_labels_present={"H"},
        )
        h = AppendixFile(
            label="H",
            data=b"%PDF-1.4",
            filename="h.pdf",
            format="pdf",
        )
        from appendix_generator import attach_appendices_to_record

        _, merged_ap, _ = attach_appendices_to_record(record, ctx, {}, [h])
        pkg = build_deliverable_zip_bytes(
            docx,
            "edge.docx",
            ctx,
            {},
            record.to_json_bytes(),
            merged_ap,
        )
        with zipfile.ZipFile(io.BytesIO(pkg)) as zf:
            names = zf.namelist()
        self.assertTrue(any(n.startswith("qp_templates/") for n in names))


class DwdaCalcResultFormatTests(unittest.TestCase):
    def test_fmt_integer_like(self) -> None:
        self.assertEqual(DwdaCalcResult._fmt(10.0), "10")

    def test_fmt_trailing_zeros_stripped(self) -> None:
        self.assertEqual(DwdaCalcResult._fmt(0.1000), "0.1")

    def test_yn_none_empty(self) -> None:
        self.assertEqual(DwdaCalcResult._yn(None), "")


if __name__ == "__main__":
    unittest.main()
